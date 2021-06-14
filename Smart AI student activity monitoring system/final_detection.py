import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler

from sklearn.linear_model import LogisticRegression, RidgeClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.metrics import accuracy_score # Accuracy metrics
import pickle
import mediapipe as mp
import cv2
import csv
import os
import time
import numpy as np
import statistics
import openpyxl
import datetime
import sys
from ctypes import cast, POINTER
from comtypes import CLSCTX_ALL
from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
wb=openpyxl.load_workbook("data.xlsx")
sheets=wb.sheetnames
sh1=wb['Sheet1']
sh1.cell(1,1,value="Time")
sh1.cell(1,2,value="P-1/A-0 ")
wb.save('data.xlsx')

devices = AudioUtilities.GetSpeakers()
interface = devices.Activate(
    IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
volume = cast(interface, POINTER(IAudioEndpointVolume))
mp_drawing = mp.solutions.drawing_utils
mp_holistic = mp.solutions.holistic
with open('smart_ai3.pkl', 'rb') as f:
    model = pickle.load(f)
ct=int(input(("How long will the class last (MINUTES) ")))
mint=int(input("AT least for how long the student should be in class (MINUTES) "))
updatet=ct/10
a=0
nop=0
nopp=0
m=0
roww=2
exitp=0
cap = cv2.VideoCapture(0)
# Initiate holistic model
with mp_holistic.Holistic(min_detection_confidence=0.75, min_tracking_confidence=0.72) as holistic:
    while cap.isOpened():
        t = str(datetime.datetime.now()).split()[1][:5]
        ret, frame = cap.read()
        image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        image.flags.writeable = False
        results = holistic.process(image)
        image.flags.writeable = True
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
        mp_drawing.draw_landmarks(image, results.face_landmarks, mp_holistic.FACE_CONNECTIONS,
                                  mp_drawing.DrawingSpec(color=(80, 110, 10), thickness=1, circle_radius=1),
                                  mp_drawing.DrawingSpec(color=(80, 256, 121), thickness=1, circle_radius=1)
                                  )
        mp_drawing.draw_landmarks(image, results.pose_landmarks, mp_holistic.POSE_CONNECTIONS,
                                  mp_drawing.DrawingSpec(color=(245, 117, 66), thickness=2, circle_radius=4),
                                  mp_drawing.DrawingSpec(color=(245, 66, 230), thickness=2, circle_radius=2)
                                  )
        a+=1
        cv2.imshow('img',image)
        cv2.waitKey(1)
        if results.face_landmarks is None and results.pose_landmarks is None and a!=(6*updatet):
            print('PLEASE ATTEND YOUR ONLINE CLASS OTHERWISE YOU WILL BE MARKED ABSENT')
            time.sleep(10)
        elif volume.GetMasterVolumeLevel()<-25 and a!=(6*updatet):
            print('INCREASE YOUR VOLUME LEVEL OTHERWISE YOU WILL BE MARKED ABSENT')
            time.sleep(10)
        else:
          try:
            if a==(6*updatet):
                    a = 0
                    if nop>=(6*updatet)/2:
                        print("ATTENDANCE UPDATE - PRESENT ")
                        flag=1
                        nop=0
                        nopp+=1
                    else:
                        print("ATTENDANCE UPDATE - ABSENT")
                        flag=0
                        nop=0
                    m += 1
                    if m%updatet==0:
                        sh1.cell(roww,1,t)
                        sh1.cell(roww,2,value=flag)
                        roww+=1
                        wb.save('data.xlsx')
                    if m-nop>ct-mint:
                        print("SORRY , YOU ARE MARKED AS ABSENT (PLEASE CONTACT YOUR SUBJECT TEACHER IF ITS WRONG)")
                        exitp=1
                        exit(1)

                    elif nopp>mint :
                        print("GREAT, YOU ARE MARKED AS PRESENT")
                        exitp = 1
                        exit(1)
            face = results.face_landmarks.landmark
            pose = results.pose_landmarks.landmark
            h, w, c = image.shape
            pose_row = list(
                np.array([[landmark.x, landmark.y, landmark.z, landmark.visibility] for landmark in pose]).flatten())
            # Extract Face landmarks
            face_row = list(
                np.array([[landmark.x, landmark.y, landmark.z, landmark.visibility] for landmark in face]).flatten())
            row = pose_row + face_row
            X = pd.DataFrame([row])
            body_language_class = model.predict(X)[0]
            if body_language_class[0]=='P':
                nop+=1
            time.sleep(10)
          except:
              if exitp==1:
                  exit(1)
